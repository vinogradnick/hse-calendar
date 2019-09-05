from datetime import datetime, date, time
import json_parser as jp
from openpyxl import Workbook
from openpyxl import load_workbook
import classes as cl
import downloader
import helper
####
MINOR_NAME = 'майнор'
COURSE = '3'
COURSE_FIRST_COLUMN = "E"


class Lesson:
    def __init__(self, name, starttime, endtime, date):
        self.name = name
        self.starttime = starttime
        self.endtime = endtime
        self.date = date

    def __str__(self):
        return (self.date.__str__() + " " + self.starttime.__str__() + " " + self.endtime.__str__() + " " + self.name.__str__())

    def __dict__(self):
        return dict(name=self.name, startTime=time_convert(self.date, self.starttime), endTime=time_convert(self.date, self.endtime))


def time_convert(dateString, timeString):
    splitedDate = [int(i) for i in dateString.split('.')]
    splitedTime = [int(i) for i in timeString.split(':')]
    datet = datetime(splitedDate[2], splitedDate[1],
                     splitedDate[0], splitedTime[0], splitedTime[1], 0)
    return datet.strftime('%Y-%m-%dT%H:%M:%S')


def get_all_lessons(file_name=downloader.FILE_NAME):
    wb = get_wb(file_name)
    sheets = get_all_sheets(wb)
    return read_sheets(sheets)


def get_wb(file_name):
    return load_workbook(filename=file_name)


def get_all_sheets(file_name=downloader.FILE_NAME):
    wb = get_wb(file_name)
    return get_all_sheets(file_name)


def get_all_sheets(wb):
    names = get_sheets_names(wb)
    sheets = get_all_sheets_by_names(wb, names)
    return sheets


def get_all_sheets_by_names(wb, sheet_names):
    sheets = []
    for sheet_name in sheet_names:
        sheets.append(cl.Page(sheet_name, get_sheet(wb, sheet_name), MINOR_NAME in sheet_name))
    return sheets


def get_sheets_names(wb):
    return ([sheet for sheet in wb.sheetnames if COURSE in sheet])


def get_sheet(wb, sheet_name):
    return wb[sheet_name]


def read_sheets(sheets):
    lessons = []
    for sheet in sheets:
        for lesson in (read_sheet(sheet)):
            lessons.append(lesson)
    return lessons


def read_sheet(page):
    lessons = []
    for i in range(4, 49):
        if page.ws["B"+i.__str__()].value != None:
            if page.is_minor:
                
            lesson = get_lesson(page, i, COURSE_FIRST_COLUMN, lessons)
            if lesson != None:
                lessons.append(lesson)
    return lessons


def get_lesson(page, excelIndex, lessonColumn, lessons):
    def validate_lesson(lesson, lessons):
        if lesson.date != None:
            lesson.date = (lesson.date.split('\n'))[1]
        if lesson.date == None:
            lesson.date = lessons[lessons.__len__()-1].date
        if lesson.name != None:
            lesson.name = lesson.name.replace('\n', ' ')
        return lesson

    time = page.ws["B"+excelIndex.__str__()].value.replace('\n',
                                                      ' ').replace('  ', ' ').split(' ')[1].split('-')
    start_time = time[0]
    end_time = time[1]

    lesson = validate_lesson(Lesson(page.ws[lessonColumn+excelIndex.__str__(
    )].value, start_time, end_time, page.ws["A"+excelIndex.__str__()].value), lessons)
    return lesson


def main():
    for x in (get_all_lessons(FILE_NAME)):
        print(jp.toJson(x.__dict__()))


if __name__ == '__main__':
    main()
